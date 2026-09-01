from rest_framework import viewsets, filters, permissions
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import AccessToken
from django.db.models import Sum, Count, Q
from django.views import View
from .models import Lead
from .serializers import LeadSerializer


class LeadViewSet(viewsets.ModelViewSet):
    serializer_class = LeadSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "contact_name", "phone_number", "email", "company_source"]
    ordering_fields = ["name", "potential_value", "created_at", "stage", "tag"]
    ordering = ["-created_at"]

    def get_queryset(self):
        user = self.request.user
        queryset = Lead.objects.filter(company=user.company)
        if user.role == "AGENT":
            queryset = queryset.filter(assigned_to=user)

        archived = self.request.query_params.get("archived")
        if archived == "true":
            queryset = queryset.filter(is_archived=True)
        else:
            queryset = queryset.filter(is_archived=False)

        stage = self.request.query_params.get("stage")
        tag = self.request.query_params.get("tag")
        assigned_to = self.request.query_params.get("assigned_to")
        if stage:
            queryset = queryset.filter(stage=stage)
        if tag:
            queryset = queryset.filter(tag=tag)
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)
        return queryset

    def get_object(self):
        pk = self.kwargs.get("pk")
        return Lead.objects.get(pk=pk)

    def perform_create(self, serializer):
        lead = serializer.save(
            company=self.request.user.company,
            assigned_to=self.request.user,
        )
        from qontak_sales.apps.notifications.models import Notification
        Notification.objects.create(
            user=self.request.user,
            title="New Lead Created",
            message=f"Lead '{lead.name}' has been added to your pipeline.",
            link=f"/leads/{lead.id}",
        )

    def destroy(self, request, *args, **kwargs):
        if request.user.role == "AGENT":
            return Response({"error": "Agents cannot delete leads"}, status=403)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"])
    def archive(self, request, pk=None):
        lead = self.get_object()
        if request.user.role == "AGENT" and lead.assigned_to != request.user:
            return Response({"error": "You can only archive your own leads"}, status=403)
        lead.is_archived = True
        lead.save()
        return Response({"message": "Lead archived", "lead": LeadSerializer(lead).data})

    @action(detail=True, methods=["post"])
    def restore(self, request, pk=None):
        lead = self.get_object()
        if request.user.role == "AGENT" and lead.assigned_to != request.user:
            return Response({"error": "You can only restore your own leads"}, status=403)
        lead.is_archived = False
        lead.save()
        return Response({"message": "Lead restored", "lead": LeadSerializer(lead).data})

    @action(detail=True, methods=["post"])
    def move_stage(self, request, pk=None):
        lead = self.get_object()
        new_stage = request.data.get("stage")
        if new_stage not in dict(Lead.STAGE_CHOICES):
            return Response({"error": "Invalid stage"}, status=400)
        old_stage = lead.get_stage_display()
        lead.stage = new_stage
        lead.save()
        from qontak_sales.apps.notifications.models import Notification
        Notification.objects.create(
            user=request.user,
            title="Lead Stage Updated",
            message=f"'{lead.name}' moved from {old_stage} to {lead.get_stage_display()}.",
            link=f"/leads/{lead.id}",
        )
        return Response(LeadSerializer(lead).data)


@api_view(["GET"])
def dashboard_stats(request):
    user = request.user
    all_company_leads = Lead.objects.filter(company=user.company, is_archived=False)

    if user.role == "AGENT":
        my_leads = all_company_leads.filter(assigned_to=user)
    else:
        my_leads = all_company_leads

    total_revenue = my_leads.filter(stage="WON").aggregate(total=Sum("potential_value"))["total"] or 0
    total_leads = my_leads.count()
    won_count = my_leads.filter(stage="WON").count()
    lost_count = my_leads.filter(stage="LOST").count()
    closed_count = won_count + lost_count
    win_rate = round((won_count / closed_count * 100) if closed_count > 0 else 0, 1)
    active_leads = my_leads.exclude(stage__in=["WON", "LOST"]).count()

    stage_dist = []
    for stage_code, stage_label in Lead.STAGE_CHOICES:
        count = my_leads.filter(stage=stage_code).count()
        stage_dist.append({"stage": stage_label, "count": count})

    monthly_data = []
    from django.utils import timezone
    from django.db.models.functions import TruncMonth
    monthly = (
        my_leads.filter(stage="WON")
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(revenue=Sum("potential_value"), count=Count("id"))
        .order_by("month")[:12]
    )
    for m in monthly:
        monthly_data.append({
            "month": m["month"].strftime("%b %Y"),
            "revenue": float(m["revenue"]),
            "count": m["count"],
        })

    from django.contrib.auth import get_user_model
    User = get_user_model()
    agents = User.objects.filter(company=user.company, role="AGENT")
    leaderboard = []
    for agent in agents:
        agent_won = all_company_leads.filter(assigned_to=agent, stage="WON")
        agent_revenue = agent_won.aggregate(total=Sum("potential_value"))["total"] or 0
        leaderboard.append({
            "name": agent.get_full_name() or agent.username,
            "deals": agent_won.count(),
            "revenue": float(agent_revenue),
        })
    leaderboard.sort(key=lambda x: x["revenue"], reverse=True)

    return Response({
        "total_revenue": float(total_revenue),
        "win_rate": win_rate,
        "active_leads": active_leads,
        "total_leads": total_leads,
        "won_count": won_count,
        "lost_count": lost_count,
        "stage_distribution": stage_dist,
        "monthly_revenue": monthly_data,
        "leaderboard": leaderboard[:10],
    })


class DashboardExportView(View):
    def get(self, request):
        import io
        from django.http import HttpResponse, JsonResponse
        from django.utils import timezone
        from django.db.models.functions import TruncMonth
        from django.contrib.auth import get_user_model
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        auth_header = request.META.get("HTTP_AUTHORIZATION", "")
        if not auth_header.startswith("Bearer "):
            return JsonResponse({"error": "Authentication required"}, status=401)
        token = auth_header.split(" ")[1]
        try:
            access_token = AccessToken(token)
            user_id = access_token["user_id"]
            User = get_user_model()
            user = User.objects.get(id=user_id)
        except Exception:
            return JsonResponse({"error": "Invalid token"}, status=401)

        all_company_leads = Lead.objects.filter(company=user.company, is_archived=False)
        my_leads = all_company_leads.filter(assigned_to=user) if user.role == "AGENT" else all_company_leads

        total_revenue = my_leads.filter(stage="WON").aggregate(total=Sum("potential_value"))["total"] or 0
        total_leads = my_leads.count()
        won_count = my_leads.filter(stage="WON").count()
        lost_count = my_leads.filter(stage="LOST").count()
        closed_count = won_count + lost_count
        win_rate = round((won_count / closed_count * 100) if closed_count > 0 else 0, 1)
        active_leads = my_leads.exclude(stage__in=["WON", "LOST"]).count()

        monthly = (
            my_leads.filter(stage="WON")
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(revenue=Sum("potential_value"), count=Count("id"))
            .order_by("month")[:12]
        )

        STAGE_MAP = dict(Lead.STAGE_CHOICES)
        TAG_MAP = dict(Lead.TAG_CHOICES)

        primary_color = "2563EB"
        header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, color=primary_color, size=16)
        subtitle_font = Font(name="Calibri", color="666666", size=10)
        section_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        label_font = Font(name="Calibri", bold=True, size=10)
        value_font = Font(name="Calibri", size=10)
        metric_value_font = Font(name="Calibri", bold=True, size=12, color=primary_color)
        zebra_light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        def style_range(ws, min_row, max_row, min_col, max_col, font=None, fill=None, alignment=None, number_format=None):
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if font: cell.font = font
                    if fill: cell.fill = fill
                    if alignment: cell.alignment = alignment
                    if number_format: cell.number_format = number_format

        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Report"
        ws.sheet_properties.tabColor = primary_color

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 22
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 18
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 16

        ws.merge_cells("A1:L1")
        ws["A1"] = "QontakSales Dashboard Report"
        ws["A1"].font = title_font
        ws["A1"].alignment = left_align

        ws.merge_cells("A2:L2")
        ws["A2"] = f"Generated: {timezone.now().strftime('%d %B %Y, %H:%M')}  |  Agent: {user.get_full_name() or user.username} ({user.role})"
        ws["A2"].font = subtitle_font

        # === SECTION 1: SUMMARY (row 4) ===
        sec_row = 4
        ws.merge_cells(f"A{sec_row}:B{sec_row}")
        ws[f"A{sec_row}"] = "SUMMARY"
        style_range(ws, sec_row, sec_row, 1, 2, font=section_font, fill=section_fill, alignment=center_align)

        summary_items = [
            ("Total Revenue", f"Rp {total_revenue:,.0f}"),
            ("Win Rate", f"{win_rate}%"),
            ("Active Leads", str(active_leads)),
            ("Total Leads", str(total_leads)),
            ("Won Deals", str(won_count)),
            ("Lost Deals", str(lost_count)),
        ]
        for i, (label, val) in enumerate(summary_items):
            r = sec_row + 1 + i
            ws.cell(row=r, column=1, value=label).font = label_font
            ws.cell(row=r, column=1).alignment = left_align
            ws.cell(row=r, column=2, value=val).font = metric_value_font if i < 2 else value_font
            ws.cell(row=r, column=2).alignment = right_align
        style_range(ws, sec_row, sec_row + len(summary_items), 1, 1, border=thin_border)
        style_range(ws, sec_row, sec_row + len(summary_items), 2, 2, border=thin_border)
        for r in range(sec_row + 1, sec_row + len(summary_items) + 1):
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2).border = thin_border
            if r % 2 == 0:
                ws.cell(row=r, column=1).fill = zebra_light
                ws.cell(row=r, column=2).fill = zebra_light

        # === SECTION 2: MONTHLY REVENUE (row 12) ===
        monthly_header_row = sec_row + len(summary_items) + 2
        ws.merge_cells(f"A{monthly_header_row}:C{monthly_header_row}")
        ws[f"A{monthly_header_row}"] = "MONTHLY REVENUE"
        style_range(ws, monthly_header_row, monthly_header_row, 1, 3, font=section_font, fill=section_fill, alignment=center_align)

        mh_row = monthly_header_row + 1
        for col_idx, h in enumerate(["Month", "Revenue", "Deals Won"], 1):
            c = ws.cell(row=mh_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        mr = mh_row + 1
        for m in monthly:
            ws.cell(row=mr, column=1, value=m["month"].strftime("%b %Y")).font = value_font
            ws.cell(row=mr, column=1).alignment = left_align
            ws.cell(row=mr, column=1).border = thin_border
            rev_cell = ws.cell(row=mr, column=2, value=float(m["revenue"]))
            rev_cell.font = value_font
            rev_cell.number_format = '"Rp "#,##0'
            rev_cell.alignment = right_align
            rev_cell.border = thin_border
            ws.cell(row=mr, column=3, value=m["count"]).font = value_font
            ws.cell(row=mr, column=3).alignment = center_align
            ws.cell(row=mr, column=3).border = thin_border
            if mr % 2 == 0:
                for c in range(1, 4):
                    ws.cell(row=mr, column=c).fill = zebra_light
            mr += 1

        if mr == mh_row + 1:
            ws.cell(row=mr, column=1, value="No data").font = Font(italic=True, color="999999")
            mr += 1

        # === SECTION 3: LEADS DATA (below monthly) ===
        leads_sec_row = mr + 1
        ws.merge_cells(f"A{leads_sec_row}:L{leads_sec_row}")
        ws[f"A{leads_sec_row}"] = "LEADS DATA"
        style_range(ws, leads_sec_row, leads_sec_row, 1, 12, font=section_font, fill=section_fill, alignment=center_align)

        leads_headers = ["Name", "Contact", "Phone", "Email", "Company", "Value", "Stage", "Tag", "Assigned To", "Created"]
        lh_row = leads_sec_row + 1
        for col_idx, h in enumerate(leads_headers, 1):
            c = ws.cell(row=lh_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        lr = lh_row + 1
        for lead in my_leads.select_related("assigned_to"):
            ws.cell(row=lr, column=1, value=lead.name).font = value_font
            ws.cell(row=lr, column=1).alignment = left_align
            ws.cell(row=lr, column=2, value=lead.contact_name).font = value_font
            ws.cell(row=lr, column=2).alignment = left_align
            ws.cell(row=lr, column=3, value=lead.phone_number).font = value_font
            ws.cell(row=lr, column=3).alignment = left_align
            ws.cell(row=lr, column=4, value=lead.email or "").font = value_font
            ws.cell(row=lr, column=4).alignment = left_align
            ws.cell(row=lr, column=5, value=lead.company_source).font = value_font
            ws.cell(row=lr, column=5).alignment = left_align
            val_cell = ws.cell(row=lr, column=6, value=float(lead.potential_value))
            val_cell.font = value_font
            val_cell.number_format = '"Rp "#,##0'
            val_cell.alignment = right_align
            ws.cell(row=lr, column=7, value=STAGE_MAP.get(lead.stage, lead.stage)).font = value_font
            ws.cell(row=lr, column=7).alignment = center_align
            ws.cell(row=lr, column=8, value=TAG_MAP.get(lead.tag, lead.tag)).font = value_font
            ws.cell(row=lr, column=8).alignment = center_align
            ws.cell(row=lr, column=9, value=lead.assigned_to.get_full_name() if lead.assigned_to else "Unassigned").font = value_font
            ws.cell(row=lr, column=9).alignment = left_align
            ws.cell(row=lr, column=10, value=lead.created_at.strftime("%d %b %Y")).font = value_font
            ws.cell(row=lr, column=10).alignment = center_align
            for c in range(1, 11):
                ws.cell(row=lr, column=c).border = thin_border
            if lr % 2 == 0:
                for c in range(1, 11):
                    ws.cell(row=lr, column=c).fill = zebra_light
            lr += 1

        if lr == lh_row + 1:
            ws.cell(row=lr, column=1, value="No leads found").font = Font(italic=True, color="999999")
            lr += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"dashboard-report-{timezone.now().strftime('%Y%m%d')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
